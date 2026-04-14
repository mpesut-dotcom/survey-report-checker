"""
Phase 0: Excel parsing — extract structured question data from XLSX files.
"""
import re
from pathlib import Path

import openpyxl

from checker.models.excel_dataset import ExcelDataset, CategoryBreakdown, DerivedMetrics


# ──────────────────────────────────────────────────────────────────────
# Constants
# ──────────────────────────────────────────────────────────────────────

_CONTENT_CODE_RE = re.compile(
    r'^(?:EXTRA\s+)?[A-Za-z]{0,8}\d+(?:x\d+)?(?:s(?:\.\d+)?)?(?:\.\d+)*[a-z]?(?:_T2B)?$',
    re.IGNORECASE,
)

_SKIP_PATTERNS = re.compile(
    r'^(\*\s*small\s+base|\*\*\s*very\s+small\s+base|significance\s+test)',
    re.IGNORECASE,
)

# ──────────────────────────────────────────────────────────────────────
# Public API
# ──────────────────────────────────────────────────────────────────────

def parse_all_excels(excel_paths: list[Path]) -> list[ExcelDataset]:
    """Parse all Excel files and return unified list of ExcelDataset."""
    all_datasets: list[ExcelDataset] = []
    for path in excel_paths:
        file_key = path.stem  # e.g. "Brand_Tracking_2024"
        datasets = parse_single_excel(path, file_key)
        all_datasets.extend(datasets)
        print(f"  {path.name}: {len(datasets)} pitanja")
    return all_datasets


def parse_single_excel(path: Path, file_key: str) -> list[ExcelDataset]:
    """Parse one Excel file into list of ExcelDataset.

    Universal approach: iterates ALL sheets, auto-detects format per sheet,
    merges regular questions + cross-tab breakdowns + MEAN values.
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    all_questions: list[dict] = []
    all_cross: dict[str, dict] = {}
    all_means: dict[str, float] = {}
    seen_q_codes: set[str] = set()

    for sname in wb.sheetnames:
        ws = wb[sname]

        # Skip TOC and tiny sheets
        if sname.upper().startswith("TOC") or ws.max_row < 3:
            continue

        # Skip significance test sheets by name pattern (_sig, _sig_total, _sig_tot)
        if re.search(r'_sig(_total|_tot)?$', sname):
            continue

        # Skip significance test sheets by content analysis (fallback)
        if _is_significance_sheet(ws):
            print(f"    Skip sig: '{sname}'")
            continue

        # Auto-detect sheet type from first header encountered
        sheet_type = _detect_sheet_type(ws)

        if sheet_type == "cross":
            # Cross-tab: extract Total-column data + means via universal parser,
            # then extract segment breakdowns via cross parser.
            questions, means = _parse_sheet_dynamic(ws)
            for q in questions:
                if q["q_code"] not in seen_q_codes:
                    seen_q_codes.add(q["q_code"])
                    all_questions.append(q)
            all_means.update(means)

            cross_data = _parse_cross_sheet(ws)
            for qc, data in cross_data.items():
                # Group by banner (sheet name) instead of flat merge
                all_cross.setdefault(qc, {})[sname] = data

            if cross_data or questions:
                print(f"    Cross '{sname}': {len(cross_data)} cross, "
                      f"{len(questions)} pitanja")
        else:
            # Regular (n|%), stats (Mean), or unknown — universal parser
            questions, means = _parse_sheet_dynamic(ws)
            for q in questions:
                if q["q_code"] not in seen_q_codes:
                    seen_q_codes.add(q["q_code"])
                    all_questions.append(q)
            all_means.update(means)
            if questions or means:
                lbl = "Regular" if sheet_type == "regular" else "Auto"
                print(f"    {lbl} '{sname}': {len(questions)} pitanja, "
                      f"{len(means)} meana")

    # Attach MEAN values to matching questions
    if all_means:
        existing_codes = {q["q_code"] for q in all_questions}
        for q in all_questions:
            if q["q_code"] in all_means and "mean_value" not in q:
                q["mean_value"] = all_means[q["q_code"]]
        for mq, mval in all_means.items():
            if mq not in existing_codes:
                all_questions.append({
                    "q_code": mq,
                    "full_text": f"{mq} (samo MEAN)",
                    "is_multioption": False,
                    "options": [],
                    "n_total": None,
                    "mean_value": mval,
                })
        print(f"    -> {len(all_means)} MEAN vrijednosti")

    # Attach cross-tab breakdowns to matching questions
    if all_cross:
        for q in all_questions:
            if q["q_code"] in all_cross:
                q["cross_tabs"] = all_cross[q["q_code"]]
        n_with_cross = sum(1 for q in all_questions if q.get("cross_tabs"))
        print(f"    -> {n_with_cross} pitanja s cross-tab podacima")

    wb.close()
    return [_to_dataset(q, file_key) for q in all_questions]


# ──────────────────────────────────────────────────────────────────────
# Conversion to new format
# ──────────────────────────────────────────────────────────────────────

def _to_dataset(q: dict, file_key: str) -> ExcelDataset:
    """Convert internal parsed dict to ExcelDataset Pydantic model."""
    q_code = q["q_code"]
    dataset_id = f"{file_key}__{q_code}"

    # Build categories with breakdowns grouped by banner
    categories: list[CategoryBreakdown] = []
    # cross_tabs structure: {banner_name: {seg_name: [{"label":..,"pct":..}], "_segment_sizes": {...}}}
    cross_tabs = q.get("cross_tabs", {})

    # Extract segment_sizes per banner
    all_segment_sizes: dict[str, dict[str, int]] = {}
    for banner_name, banner_data in cross_tabs.items():
        raw_sizes = banner_data.get("_segment_sizes", {})
        if raw_sizes:
            all_segment_sizes[banner_name] = {}
            for seg, n in raw_sizes.items():
                try:
                    all_segment_sizes[banner_name][seg] = int(round(float(n)))
                except (ValueError, TypeError):
                    pass

    for opt in q.get("options", []):
        label = opt["label"]
        total = opt["total_pct"]

        # Find breakdowns for this option, grouped by banner
        breakdowns: dict[str, dict[str, float | None]] = {}
        for banner_name, banner_data in cross_tabs.items():
            for seg_name, seg_entries in banner_data.items():
                if seg_name == "_segment_sizes":
                    continue
                # seg_entries is list of {"label": ..., "pct": ...}
                for entry in seg_entries:
                    if entry["label"] == label:
                        breakdowns.setdefault(banner_name, {})[seg_name] = entry["pct"]
                        break

        categories.append(CategoryBreakdown(
            label=label,
            total=total,
            breakdowns=breakdowns,
        ))

    # Derived metrics
    derived = DerivedMetrics()
    if q.get("mean_value") is not None:
        derived.mean = q["mean_value"]

    # Auto-detect top2box for scale questions
    if len(categories) >= 4:
        labels_lower = [c.label.lower() for c in categories]
        # Heuristic: if labels look like a scale (1-5, very ... not at all, etc.)
        if _looks_like_scale(labels_lower) and categories[0].total is not None and categories[1].total is not None:
            derived.top2box = round(categories[0].total + categories[1].total, 1)

    # Detect question type
    q_type = _detect_type(q, categories)

    return ExcelDataset(
        id=dataset_id,
        file_key=file_key,
        sheet_name=q_code,
        question_code=q_code,
        question_text=q["full_text"],
        base_description=None,
        base_n=q.get("n_total"),
        type=q_type,
        categories=categories,
        derived_metrics=derived if (derived.mean is not None or derived.top2box is not None or derived.net is not None) else None,
        segment_sizes=all_segment_sizes,
    )


def _detect_type(q: dict, categories: list[CategoryBreakdown]) -> str:
    """Detect question type from data."""
    if q.get("is_multioption"):
        return "multi_choice"
    if q.get("mean_value") is not None:
        return "scale"
    if not categories:
        return "open"
    # Check if totals sum to ~100%
    totals = [c.total for c in categories if c.total is not None]
    if totals:
        s = sum(totals)
        if 95 <= s <= 105:
            return "single_choice"
        if s > 110:
            return "multi_choice"
    return "single_choice"


def _looks_like_scale(labels: list[str]) -> bool:
    """Heuristic check if labels represent a scale."""
    scale_words = {"vrlo", "uglavnom", "donekle", "malo", "nimalo",
                   "potpuno", "djelomično", "definitely", "probably",
                   "somewhat", "not", "1", "2", "3", "4", "5",
                   "izrazito", "prilično", "osrednje", "nezadovoljan",
                   "zadovoljan", "odlično", "dobro", "loše"}
    hits = sum(1 for lbl in labels if any(w in lbl for w in scale_words))
    return hits >= 2


# ──────────────────────────────────────────────────────────────────────
# Core sheet parsing (proven v1 logic)
# ──────────────────────────────────────────────────────────────────────

def _detect_header(ws, row: int, max_col: int) -> dict | None:
    """Detect table header row format."""
    a_val = ws.cell(row, 1).value
    if a_val is not None:
        return None

    b_val = ws.cell(row, 2).value
    if b_val is None:
        return None

    b_str = str(b_val).strip()
    c_val = ws.cell(row, 3).value

    # Format B: n | %
    if b_str.lower() == "n" and c_val is not None and str(c_val).strip() == "%":
        return {"type": "n_pct", "pct_col": 3, "n_col": 2}

    # Format A: Total
    if b_str.lower() == "total":
        return {"type": "total", "pct_col": 2, "n_col": None}

    # Format C: Mean | Std.Dev. | ... | N (stats tables)
    if b_str.lower() == "mean":
        n_col = None
        for c in range(3, min(max_col + 1, 15)):
            v = ws.cell(row, c).value
            if v is not None and str(v).strip().upper() == "N":
                n_col = c
                break
        return {"type": "stats", "pct_col": 2, "n_col": n_col}

    return None


def _is_sig_letter_row(ws, row: int, max_col: int) -> bool:
    """Detect significance test letter row."""
    a_val = ws.cell(row, 1).value
    if a_val is not None:
        return False
    letters = 0
    for c in range(2, min(max_col + 1, 10)):
        v = ws.cell(row, c).value
        if v is not None and re.match(r'^[A-Z]$', str(v).strip()):
            letters += 1
    return letters >= 2


def _parse_sheet_dynamic(ws) -> tuple[list[dict], dict[str, float]]:
    """
    Universal parser that dynamically detects table format.
    Returns (questions_list, mean_values_dict).
    """
    questions: list[dict] = []
    mean_values: dict[str, float] = {}
    seen_codes: set[str] = set()
    max_row = ws.max_row
    max_col = ws.max_column
    row = 1

    while row <= max_row:
        cell_a = ws.cell(row, 1).value
        cell_b = ws.cell(row, 2).value

        if cell_a is not None and cell_b is None:
            text = str(cell_a).strip()

            if not text or text == "*Multioption" or _SKIP_PATTERNS.match(text):
                row += 1
                continue

            text_clean = re.sub(r'\s*\(Table\s+[\d.]+[^)]*\)\s*$', '', text).strip()

            # MEAN tables
            is_mean_table = "- MEAN" in text
            if is_mean_table:
                q_match_mean = re.match(
                    r'^((?:EXTRA\s+)?[A-Za-z]*\d+(?:x\d+)?(?:s(?:\.\d+)?)?(?:\.\d+)?[a-z]?)', text)
                mean_q_code = q_match_mean.group(1) if q_match_mean else None
                if mean_q_code:
                    next_row_m = row + 1
                    if next_row_m <= max_row:
                        header_m = _detect_header(ws, next_row_m, max_col)
                        if header_m:
                            pct_col_m = header_m["pct_col"]
                            data_row_m = next_row_m + 1
                            if data_row_m <= max_row and _is_sig_letter_row(ws, data_row_m, max_col):
                                data_row_m += 1
                            scan = data_row_m
                            while scan <= max_row and scan < data_row_m + 10:
                                lbl = ws.cell(scan, 1).value
                                if lbl is None:
                                    break
                                lbl_s = str(lbl).strip().lower()
                                if lbl_s in ("mean", "prosjek", "prosječna ocjena", "average"):
                                    val = ws.cell(scan, pct_col_m).value
                                    if val is not None:
                                        try:
                                            mean_values[mean_q_code] = round(float(val), 4)
                                        except (ValueError, TypeError):
                                            pass
                                    break
                                scan += 1
                row += 1
                continue

            # Look for header row after title
            next_row = row + 1
            if next_row > max_row:
                row += 1
                continue

            header = _detect_header(ws, next_row, max_col)
            if header is None:
                row += 1
                continue

            pct_col = header["pct_col"]
            n_col = header["n_col"]

            is_t2b = text.endswith("- T2B")
            q_match = re.match(
                r'^((?:EXTRA\s+)?[A-Za-z]*\d+(?:x\d+)?(?:s(?:\.\d+)?)?(?:\.\d+)?[a-z]?)', text)
            q_code = q_match.group(1) if q_match else text_clean[:80]
            if is_t2b:
                q_code = q_code + "_T2B"

            full_text = text_clean

            if q_code in seen_codes:
                row += 1
                continue

            # Skip to data rows
            data_row = next_row + 1
            if data_row <= max_row and _is_sig_letter_row(ws, data_row, max_col):
                data_row += 1

            # Read options
            row = data_row
            options = []
            n_total = None
            is_multioption = False

            while row <= max_row:
                opt_a = ws.cell(row, 1).value
                opt_pct = ws.cell(row, pct_col).value

                all_none = all(ws.cell(row, c).value is None for c in range(1, min(max_col + 1, 5)))
                if all_none:
                    break

                label = str(opt_a).strip() if opt_a else ""

                if label == "N":
                    try:
                        n_total = int(round(float(ws.cell(row, pct_col).value)))
                    except (ValueError, TypeError):
                        n_total = None
                    row += 1
                    break

                if label.startswith("Total"):
                    if n_col is not None:
                        try:
                            n_total = int(round(float(ws.cell(row, n_col).value)))
                        except (ValueError, TypeError):
                            n_total = None
                    if label == "Total*":
                        is_multioption = True
                    row += 1
                    break

                if label == "*Multioption":
                    is_multioption = True
                    row += 1
                    break

                pct = opt_pct
                if pct is not None:
                    try:
                        pct = round(float(pct), 1)
                    except (ValueError, TypeError):
                        pct = None
                options.append({"label": label, "total_pct": pct})
                row += 1

            # Check *Multioption after last row
            if row <= max_row:
                peek = ws.cell(row, 1).value
                if peek and str(peek).strip() == "*Multioption":
                    is_multioption = True

            # Stats headers: read N from data row column, set mean_value
            if header["type"] == "stats":
                if n_total is None and n_col is not None:
                    nv = ws.cell(data_row, n_col).value
                    if nv is not None:
                        try:
                            n_total = int(round(float(nv)))
                        except (ValueError, TypeError):
                            pass
                if options and options[0]["total_pct"] is not None:
                    mean_values[q_code] = options[0]["total_pct"]

            if options:
                seen_codes.add(q_code)
                q_dict = {
                    "q_code": q_code,
                    "full_text": full_text,
                    "is_multioption": is_multioption,
                    "options": options,
                    "n_total": n_total,
                }
                if header["type"] == "stats" and options:
                    first_mean = next(
                        (o["total_pct"] for o in options if o["total_pct"] is not None),
                        None,
                    )
                    if first_mean is not None:
                        q_dict["mean_value"] = first_mean
                questions.append(q_dict)
            continue
        row += 1

    return questions, mean_values


def _parse_cross_sheet(ws) -> dict[str, dict]:
    """
    Parse CROSS sheet for cross-tab (breakdown) data.
    Returns: {q_code: {segment_name: [{"label": ..., "pct": ...}], "_segment_sizes": {...}}}
    """
    result: dict[str, dict] = {}
    seen_codes: set[str] = set()
    max_row = ws.max_row
    max_col = ws.max_column
    row = 1

    while row <= max_row:
        cell_a = ws.cell(row, 1).value
        cell_b = ws.cell(row, 2).value

        if cell_a is not None and cell_b is None:
            text = str(cell_a).strip()

            if not text or text == "*Multioption" or _SKIP_PATTERNS.match(text):
                row += 1
                continue

            text_clean = re.sub(r'\s*\(Table\s+[\d.]+[^)]*\)\s*$', '', text).strip()

            if "- MEAN" in text:
                row += 1
                continue

            next_row = row + 1
            if next_row > max_row:
                row += 1
                continue

            a_hdr = ws.cell(next_row, 1).value
            b_hdr = ws.cell(next_row, 2).value
            if a_hdr is not None or b_hdr is None or str(b_hdr).strip().lower() != "total":
                row += 1
                continue

            segments: dict[int, str] = {}
            for c in range(3, max_col + 1):
                v = ws.cell(next_row, c).value
                if v is not None:
                    segments[c] = str(v).strip()

            if not segments:
                row += 1
                continue

            q_match = re.match(
                r'^((?:EXTRA\s+)?[A-Za-z]*\d+(?:x\d+)?(?:s(?:\.\d+)?)?(?:\.\d+)?[a-z]?)', text)
            q_code = q_match.group(1) if q_match else None
            if not q_code or q_code in seen_codes:
                row += 1
                continue

            data_row = next_row + 1
            if data_row <= max_row and _is_sig_letter_row(ws, data_row, max_col):
                data_row += 1

            cross_data: dict[str, list[dict]] = {seg_name: [] for seg_name in segments.values()}
            segment_n: dict[str, int] = {}
            row = data_row

            while row <= max_row:
                opt_a = ws.cell(row, 1).value
                all_none = all(ws.cell(row, c).value is None for c in range(1, min(max_col + 1, 5)))
                if all_none:
                    break

                label = str(opt_a).strip() if opt_a else ""

                if label == "N":
                    for col, seg_name in segments.items():
                        nv = ws.cell(row, col).value
                        if nv is not None:
                            try:
                                segment_n[seg_name] = int(round(float(nv)))
                            except (ValueError, TypeError):
                                pass
                    row += 1
                    break

                if label in ("Total", "Total*", "*Multioption"):
                    row += 1
                    break

                for col, seg_name in segments.items():
                    v = ws.cell(row, col).value
                    pct = None
                    if v is not None:
                        try:
                            pct = round(float(v), 1)
                        except (ValueError, TypeError):
                            pct = None
                    cross_data[seg_name].append({"label": label, "pct": pct})
                row += 1

            if any(cross_data.values()):
                seen_codes.add(q_code)
                if segment_n:
                    cross_data["_segment_sizes"] = segment_n
                result[q_code] = cross_data
            continue
        row += 1

    return result


# ──────────────────────────────────────────────────────────────────────
# Sheet type detection
# ──────────────────────────────────────────────────────────────────────

def _detect_sheet_type(ws) -> str:
    """Detect sheet format: 'regular' (n|%), 'cross' (Total+segments), or 'unknown'.

    Scans first ~50 rows for the first recognizable header row.
    """
    for row in range(1, min(50, ws.max_row + 1)):
        a = ws.cell(row, 1).value
        b = ws.cell(row, 2).value

        # Header rows have A=None, B=keyword
        if a is not None or b is None:
            continue

        b_str = str(b).strip().lower()

        if b_str == "n":
            c = ws.cell(row, 3).value
            if c is not None and str(c).strip() == "%":
                return "regular"

        elif b_str == "total":
            # Total with segment columns → cross; Total alone → regular
            c = ws.cell(row, 3).value
            if c is not None and str(c).strip():
                return "cross"
            return "regular"

        elif b_str == "mean":
            return "regular"  # stats tables handled by _parse_sheet_dynamic

    return "unknown"


def _is_significance_sheet(ws) -> bool:
    """Detect significance test sheets where data cells contain letters (A, AB, BCD)
    instead of numeric values.

    These sheets look like cross-tabs but their data cells hold significance
    markers rather than numbers.
    """
    max_col = ws.max_column

    # Find the first header-like row
    header_row = None
    for r in range(1, min(30, ws.max_row + 1)):
        a = ws.cell(r, 1).value
        b = ws.cell(r, 2).value
        if a is None and b is not None and str(b).strip().lower() in ("total", "n"):
            header_row = r
            break

    if not header_row:
        return False

    # Skip the letter-assignment row (A, B, C labels — normal in all cross-tabs)
    data_start = header_row + 1
    if data_start <= ws.max_row and _is_sig_letter_row(ws, data_start, max_col):
        data_start += 1

    # Sample actual data cells and check if they're mostly letters/dashes vs numbers
    numeric_cells = 0
    letter_cells = 0

    for r in range(data_start, min(data_start + 20, ws.max_row + 1)):
        label = ws.cell(r, 1).value
        if label is not None and str(label).strip() in ("N", "Total", "Total*", "*Multioption"):
            break

        for c in range(2, min(max_col + 1, 10)):
            v = ws.cell(r, c).value
            if v is None:
                continue
            if isinstance(v, (int, float)):
                numeric_cells += 1
            else:
                s = str(v).strip()
                if re.match(r'^[A-Z]{1,8}$', s) or s == "-":
                    letter_cells += 1
                else:
                    try:
                        float(s)
                        numeric_cells += 1
                    except ValueError:
                        letter_cells += 1

    total = numeric_cells + letter_cells
    if total < 5:
        return False
    # Significance sheets have more letter/dash cells than numeric ones
    return letter_cells > numeric_cells
